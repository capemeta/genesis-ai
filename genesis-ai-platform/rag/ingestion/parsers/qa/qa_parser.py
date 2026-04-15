"""
QA 结构化解析器

第一期仅支持固定模板的 CSV / Excel 问答导入：
- .csv
- .xlsx
"""

import csv
import logging
from io import BytesIO
from io import StringIO
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook  # type: ignore[import-untyped]

from rag.ingestion.parsers.base import BaseParser
from rag.ingestion.parsers.encoding_utils import decode_with_encoding_detection
from utils.qa_markdown import build_qa_markdown_text

logger = logging.getLogger(__name__)


class QAParser(BaseParser):
    """
    QA 结构化解析器。

    为了降低运营复杂度，一期严格使用固定模板列，不做动态字段映射。
    """

    SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
    REQUIRED_HEADERS = {"question", "answer"}
    OPTIONAL_HEADERS = {"similar_questions", "category", "tags", "enabled"}
    ALL_HEADERS = REQUIRED_HEADERS | OPTIONAL_HEADERS

    def supports(self, file_extension: str) -> bool:
        return file_extension.lower() in self.SUPPORTED_EXTENSIONS

    def parse(self, file_buffer: bytes, file_extension: str) -> Tuple[str, Dict[str, Any]]:
        """解析固定模板的 QA 文件并输出统一 qa_items 结构。"""
        ext = file_extension.lower()
        logger.info("[QAParser] 开始解析 QA 数据，格式: %s", ext)

        if ext == ".csv":
            qa_items = self._parse_csv(file_buffer)
        elif ext == ".xlsx":
            qa_items = self._parse_excel(file_buffer)
        else:
            raise ValueError(f"QAParser 不支持的文件类型: {file_extension}，当前仅支持 .csv / .xlsx")

        if not qa_items:
            raise ValueError("未解析到有效问答，请检查 question 和 answer 列是否填写完整")

        preview_text = self._build_preview_markdown(qa_items)
        metadata = {
            "parse_method": "qa",
            "parser": "qa",
            "qa_items": qa_items,
            "qa_item_count": len(qa_items),
            "element_count": len(qa_items),
            "template_version": "qa_import_v1",
            "supported_file_types": [".csv", ".xlsx"],
        }
        return preview_text, metadata

    def _parse_csv(self, file_buffer: bytes) -> List[Dict[str, Any]]:
        """解析 CSV，支持 UTF-8 / UTF-8-SIG 等常见编码。"""
        text, _ = decode_with_encoding_detection(file_buffer)
        sample = "\n".join(text.splitlines()[:5])
        try:
            delimiter = csv.Sniffer().sniff(sample).delimiter
        except Exception:
            delimiter = ","

        reader = csv.DictReader(StringIO(text), delimiter=delimiter)
        fieldnames = [self._normalize_header_value(name) for name in (reader.fieldnames or [])]
        self._validate_headers(fieldnames)

        records: List[Dict[str, Any]] = []
        for row_index, row in enumerate(reader, start=2):
            record: Dict[str, Any] = {
                self._normalize_header_value(key): ("" if value is None else str(value).strip())
                for key, value in dict(row).items()
                if self._normalize_header_value(key)
            }
            if any(str(v).strip() for v in record.values()):
                record["_row_index"] = row_index
                records.append(record)
        return self._normalize_records(records, source_name="csv")

    def _parse_excel(self, file_buffer: bytes) -> List[Dict[str, Any]]:
        """解析 Excel，第一阶段只读取第一个 sheet。"""
        workbook = load_workbook(BytesIO(file_buffer), data_only=True, read_only=True)
        if not workbook.worksheets:
            return []

        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        header = [self._normalize_header_value(v) for v in rows[0]]
        self._validate_headers(header)

        records: List[Dict[str, Any]] = []
        for row_index, row in enumerate(rows[1:], start=2):
            record: Dict[str, Any] = {}
            for col_idx, header_name in enumerate(header):
                if not header_name:
                    continue
                cell = row[col_idx] if col_idx < len(row) else None
                record[header_name] = "" if cell is None else str(cell).strip()

            if any(str(v).strip() for v in record.values()):
                record["_sheet_name"] = sheet.title
                record["_row_index"] = row_index
                records.append(record)

        return self._normalize_records(records, source_name="excel")

    def _validate_headers(self, headers: List[str]) -> None:
        """校验是否符合固定模板列要求。"""
        normalized_headers = [header for header in headers if header]
        header_set = set(normalized_headers)
        missing_headers = sorted(self.REQUIRED_HEADERS - header_set)
        if missing_headers:
            raise ValueError(f"QA 导入模板缺少必填列: {', '.join(missing_headers)}")

        unknown_headers = sorted(header_set - self.ALL_HEADERS)
        if unknown_headers:
            raise ValueError(
                "QA 导入模板存在未定义列: "
                f"{', '.join(unknown_headers)}。"
                "请使用固定模板列：question, answer, similar_questions, category, tags, enabled"
            )

    def _normalize_records(self, records: List[Dict[str, Any]], source_name: str) -> List[Dict[str, Any]]:
        """将导入记录规范化为统一 qa_items 协议。"""
        normalized: List[Dict[str, Any]] = []
        for idx, raw in enumerate(records, start=1):
            question = str((raw or {}).get("question") or "").strip()
            answer = str((raw or {}).get("answer") or "").strip()
            if not question or not answer:
                continue

            similar_questions = self._split_similar_questions(str((raw or {}).get("similar_questions") or ""))
            tags = self._split_tags(str((raw or {}).get("tags") or ""))
            category = str((raw or {}).get("category") or "").strip() or None
            is_enabled = self._parse_enabled_value((raw or {}).get("enabled"))

            normalized.append(
                {
                    "record_id": f"{source_name}-{idx}",
                    "question": question,
                    "answer": answer,
                    "similar_questions": similar_questions,
                    "tags": tags,
                    "category": category,
                    "source_row": int(raw.get("_row_index") or idx),
                    "source_sheet_name": raw.get("_sheet_name"),
                    "is_enabled": is_enabled,
                }
            )
        self._validate_duplicate_questions(normalized)
        return normalized

    @staticmethod
    def _normalize_header_value(value: Any) -> str:
        """规范化表头名称，统一按小写英文列名处理。"""
        return "" if value is None else str(value).strip().lower()

    @staticmethod
    def _split_similar_questions(value: str) -> List[str]:
        """按固定模板规则拆分相似问题。"""
        if not value:
            return []

        deduped: List[str] = []
        seen = set()
        for part in value.split("||"):
            text = str(part).strip()
            if not text or text in seen:
                continue
            seen.add(text)
            deduped.append(text)
        return deduped

    @staticmethod
    def _split_tags(value: str) -> List[str]:
        """按标签字段的宽松分隔规则拆分标签。"""
        if not value:
            return []

        separators = ["，", ",", "；", ";", "\n"]
        parts = [value]
        for separator in separators:
            next_parts: List[str] = []
            for part in parts:
                next_parts.extend(part.split(separator))
            parts = next_parts

        deduped: List[str] = []
        seen = set()
        for part in parts:
            text = str(part).strip()
            if not text or text in seen:
                continue
            seen.add(text)
            deduped.append(text)
        return deduped

    @staticmethod
    def _parse_enabled_value(value: Any) -> bool:
        """解析启停字段，未填写时默认启用。"""
        if value is None:
            return True

        text = str(value).strip().lower()
        if not text:
            return True

        truthy = {"true", "1", "yes", "y", "是", "启用"}
        falsy = {"false", "0", "no", "n", "否", "禁用"}

        if text in truthy:
            return True
        if text in falsy:
            return False
        raise ValueError(f"enabled 列存在非法值: {value}")

    @staticmethod
    def _validate_duplicate_questions(qa_items: List[Dict[str, Any]]) -> None:
        """校验模板中是否存在重复问题，避免导入出多条语义相同的记录。"""
        question_rows: Dict[str, List[int]] = {}
        for item in qa_items:
            question = str(item.get("question") or "").strip()
            source_row = int(item.get("source_row") or 0)
            if not question:
                continue
            question_rows.setdefault(question, []).append(source_row)

        duplicated = {
            question: rows
            for question, rows in question_rows.items()
            if len(rows) > 1
        }
        if not duplicated:
            return

        first_question, first_rows = next(iter(duplicated.items()))
        raise ValueError(
            f"检测到重复问题：{first_question}；重复行号：{', '.join(str(row) for row in first_rows)}"
        )

    @staticmethod
    def _build_preview_markdown(qa_items: List[Dict[str, Any]]) -> str:
        """构造简易预览文本。"""
        parts: List[str] = []
        for item in qa_items:
            parts.append(
                build_qa_markdown_text(
                    question=str(item.get("question") or "").strip(),
                    answer=str(item.get("answer") or "").strip(),
                    similar_questions=item.get("similar_questions") or [],
                    category=str(item.get("category") or "").strip(),
                    tags=item.get("tags") or [],
                )
            )
            parts.append("")
        return "\n".join(parts).strip()
