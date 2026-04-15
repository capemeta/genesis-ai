"""
Excel 通用模式解析器

支持 .xlsx / .xls 格式，输出：
- text: 整表 Markdown 字符串（供展示）
- metadata: 含 sheets_data 结构化行数据（供 ExcelGeneralChunker 直接读取，无需重新解析 Markdown）

改进点（相对旧版）：
1. 按文件大小选择 read_only 模式（<=50MB 用 full load 支持合并单元格）
2. 智能表头检测（扫描前 N 行，参考 Dify）
3. 合并单元格回填（full load 模式，参考 MaxKB/KnowFlow）
4. 超链接转 Markdown [text](url)（参考 Dify）
5. 日期/错误/None 单元格规范化（参考 UltraRAG）
6. 隐藏 Sheet 跳过（可配置，参考 UltraRAG）
7. openpyxl → pandas 双重兜底（参考 KnowFlow）
8. 列类型推断，写入 metadata
9. metadata 新增 sheets_data 结构化字段（供 ExcelGeneralChunker 消费）
"""

import logging
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from .excel_parser_utils import (
    backfill_merged_cells,
    extract_hyperlink,
    find_header_row,
    infer_column_types,
    normalize_cell_value,
    rows_to_markdown,
)

logger = logging.getLogger(__name__)

# 文件大小阈值：超过此值使用 read_only 流式读取（但无法处理合并单元格）
FULL_LOAD_SIZE_LIMIT = 50 * 1024 * 1024  # 50MB


class ExcelParser:
    """
    Excel 通用模式解析器。

    将 Excel 工作簿解析为 Markdown 文本（供展示）以及结构化行数据（供分块器消费）。
    两种输出均存放在返回的 (text, metadata) 中：
        - text: 各 Sheet 的 Markdown 表格拼接（以 "\\n\\n---\\n\\n" 分隔）
        - metadata["sheets_data"]: 结构化行数据列表，ExcelGeneralChunker 优先从这里读取
    """

    def __init__(
        self,
        max_rows: int = 10000,
        max_cols: int = 100,
        scan_header_rows: int = 10,
        include_hidden_sheets: bool = False,
    ):
        """
        Args:
            max_rows: 每个工作表最大行数（防止内存溢出）
            max_cols: 每个工作表最大列数
            scan_header_rows: 智能表头检测扫描的行数
            include_hidden_sheets: 是否处理隐藏工作表（默认跳过）
        """
        self.max_rows = max_rows
        self.max_cols = max_cols
        self.scan_header_rows = scan_header_rows
        self.include_hidden_sheets = include_hidden_sheets

    def is_available(self) -> bool:
        """检查 openpyxl 是否可用。"""
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            logger.warning("[ExcelParser] openpyxl 未安装")
            return False

    def parse(
        self,
        file_buffer: bytes,
        file_extension: str = ".xlsx",
    ) -> Tuple[str, Dict[str, Any]]:
        """
        解析 Excel 文档。

        Args:
            file_buffer: Excel 文件字节内容
            file_extension: 文件扩展名（.xlsx 或 .xls）

        Returns:
            (markdown_text, metadata): Markdown 文本和元数据
        """
        logger.info(f"[ExcelParser] 开始解析，格式: {file_extension}, 大小: {len(file_buffer)} bytes")

        if file_extension.lower() == ".xls":
            return self._parse_xls(file_buffer)
        else:
            return self._parse_xlsx(file_buffer)

    # ------------------------------------------------------------------ #
    # .xlsx 解析
    # ------------------------------------------------------------------ #

    def _parse_xlsx(self, file_buffer: bytes) -> Tuple[str, Dict[str, Any]]:
        """解析 .xlsx 文件，openpyxl 失败时降级 pandas。"""
        try:
            return self._parse_xlsx_openpyxl(file_buffer)
        except Exception as e:
            logger.warning(f"[ExcelParser] openpyxl 解析失败，降级 pandas: {e}")
            return self._parse_xlsx_pandas(file_buffer)

    def _parse_xlsx_openpyxl(self, file_buffer: bytes) -> Tuple[str, Dict[str, Any]]:
        """使用 openpyxl 解析 .xlsx 文件。"""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl 未安装，请执行: pip install openpyxl")

        file_size = len(file_buffer)
        use_read_only = file_size > FULL_LOAD_SIZE_LIMIT

        if use_read_only:
            logger.info(f"[ExcelParser] 文件 {file_size/1024/1024:.1f}MB > 50MB，使用 read_only 流式读取（跳过合并单元格处理）")
            wb = openpyxl.load_workbook(BytesIO(file_buffer), read_only=True, data_only=True, keep_vba=False)
        else:
            wb = openpyxl.load_workbook(BytesIO(file_buffer), read_only=False, data_only=True, keep_vba=False)

        sheets_markdown: List[str] = []
        sheets_data: List[Dict[str, Any]] = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # 跳过隐藏 Sheet
            if not self.include_hidden_sheets:
                state = getattr(sheet, "sheet_state", None)
                if state == "hidden":
                    logger.debug(f"[ExcelParser] 跳过隐藏 Sheet: {sheet_name}")
                    continue

            # 合并单元格回填（仅 full load 模式可用）
            if not use_read_only:
                backfill_merged_cells(sheet)
                has_merged = bool(list(sheet.merged_cells.ranges)) if hasattr(sheet, "merged_cells") else False
            else:
                has_merged = False

            # 读取所有行（含原始值，供表头检测用）
            raw_rows: List[List[Any]] = []
            formula_none_count = 0

            for row_idx, row in enumerate(sheet.iter_rows(values_only=False)):
                if row_idx >= self.max_rows:
                    logger.warning(f"[ExcelParser] Sheet '{sheet_name}' 行数超过 {self.max_rows}，截断")
                    break
                row_data: List[Any] = []
                for col_idx, cell in enumerate(row):
                    if col_idx >= self.max_cols:
                        break
                    # 超链接处理（full load 模式）
                    if not use_read_only:
                        hl = extract_hyperlink(cell)
                        if hl is not None:
                            row_data.append(hl)
                            continue
                    # 公式缓存 None 计数
                    if cell.value is None and hasattr(cell, "data_type") and cell.data_type == "f":
                        formula_none_count += 1
                    row_data.append(cell.value)
                raw_rows.append(row_data)

            if not raw_rows:
                logger.info(f"[ExcelParser] Sheet '{sheet_name}' 为空，跳过")
                continue

            # 智能表头检测
            header_row_idx = find_header_row(raw_rows, self.scan_header_rows)

            # 表头行和数据行分离
            header_raw = raw_rows[header_row_idx]
            data_raw = raw_rows[header_row_idx + 1:]

            # 规范化表头
            header = [normalize_cell_value(v) or f"列{i+1}" for i, v in enumerate(header_raw[:self.max_cols])]

            # 规范化数据行
            str_rows: List[List[str]] = []
            for row in data_raw:
                str_row = [normalize_cell_value(v) for v in row[:len(header)]]
                # 补齐列数
                str_row += [""] * max(0, len(header) - len(str_row))
                str_rows.append(str_row)

            # 过滤全空行
            str_rows = [r for r in str_rows if any(v for v in r)]

            if not str_rows:
                logger.info(f"[ExcelParser] Sheet '{sheet_name}' 无有效数据行，跳过")
                continue

            # 列类型推断
            column_types = infer_column_types(header, str_rows)

            # 公式缓存告警
            if formula_none_count > len(str_rows) * 0.1:
                logger.warning(
                    f"[ExcelParser] Sheet '{sheet_name}' 有 {formula_none_count} 个公式单元格无缓存值"
                    "（建议先在 Excel 中打开保存后再上传）"
                )

            # 生成 Markdown（供展示）
            sheet_md = rows_to_markdown(sheet_name, header, str_rows, row_start=1)
            if sheet_md:
                sheets_markdown.append(sheet_md)

            # 结构化数据（供 ExcelGeneralChunker 消费）
            sheets_data.append({
                "sheet_name": sheet_name,
                "header": header,
                "header_row_index": header_row_idx,
                "rows": str_rows,
                "column_types": column_types,
                "has_merged_cells": has_merged,
                "formula_none_cells": formula_none_count,
            })

        markdown = "\n\n---\n\n".join(sheets_markdown)
        metadata = {
            "parse_method": "excel",
            "parser": "openpyxl" if not (len(file_buffer) > FULL_LOAD_SIZE_LIMIT) else "openpyxl_readonly",
            "format": "xlsx",
            "sheet_count": len(sheets_data),
            "sheets": [{"name": s["sheet_name"], "rows": len(s["rows"]), "cols": len(s["header"])} for s in sheets_data],
            "sheets_data": sheets_data,
        }

        logger.info(f"[ExcelParser] xlsx 解析完成，有效 Sheet 数: {len(sheets_data)}")
        return markdown, metadata

    def _parse_xlsx_pandas(self, file_buffer: bytes) -> Tuple[str, Dict[str, Any]]:
        """使用 pandas 解析 .xlsx（openpyxl 失败时的降级路径）。"""
        try:
            import pandas as pd
        except ImportError:
            raise RuntimeError("pandas 未安装，请执行: pip install pandas openpyxl")

        xls = pd.ExcelFile(BytesIO(file_buffer), engine="openpyxl")
        sheets_markdown: List[str] = []
        sheets_data: List[Dict[str, Any]] = []

        for sheet_name in xls.sheet_names:
            df = xls.parse(sheet_name, header=None, dtype=str, nrows=self.max_rows)
            if df.empty:
                continue

            raw_rows = df.values.tolist()
            header_row_idx = find_header_row(raw_rows, self.scan_header_rows)

            header_raw = raw_rows[header_row_idx]
            data_raw = raw_rows[header_row_idx + 1:]

            header = [str(v).strip() if v and str(v).strip() not in ("nan", "None") else f"列{i+1}"
                      for i, v in enumerate(header_raw[:self.max_cols])]

            str_rows: List[List[str]] = []
            for row in data_raw:
                str_row = [str(v).strip() if v and str(v).strip() not in ("nan", "None") else ""
                           for v in row[:len(header)]]
                str_row += [""] * max(0, len(header) - len(str_row))
                str_rows.append(str_row)
            str_rows = [r for r in str_rows if any(v for v in r)]

            if not str_rows:
                continue

            column_types = infer_column_types(header, str_rows)
            sheet_md = rows_to_markdown(sheet_name, header, str_rows, row_start=1)
            if sheet_md:
                sheets_markdown.append(sheet_md)

            sheets_data.append({
                "sheet_name": sheet_name,
                "header": header,
                "header_row_index": header_row_idx,
                "rows": str_rows,
                "column_types": column_types,
                "has_merged_cells": False,
                "formula_none_cells": 0,
            })

        markdown = "\n\n---\n\n".join(sheets_markdown)
        metadata = {
            "parse_method": "excel",
            "parser": "pandas_fallback",
            "format": "xlsx",
            "sheet_count": len(sheets_data),
            "sheets": [{"name": s["sheet_name"], "rows": len(s["rows"]), "cols": len(s["header"])} for s in sheets_data],
            "sheets_data": sheets_data,
        }

        logger.info(f"[ExcelParser] xlsx 降级 pandas 解析完成，有效 Sheet 数: {len(sheets_data)}")
        return markdown, metadata

    # ------------------------------------------------------------------ #
    # .xls 解析
    # ------------------------------------------------------------------ #

    def _parse_xls(self, file_buffer: bytes) -> Tuple[str, Dict[str, Any]]:
        """解析 .xls 文件（旧版二进制格式）。"""
        try:
            import xlrd
        except ImportError:
            raise RuntimeError("xlrd 未安装，请执行: pip install xlrd")

        # on_demand=True：按需加载 Sheet，减少大文件内存占用（参考 UltraRAG）
        wb = xlrd.open_workbook(file_contents=file_buffer, on_demand=True)

        sheets_markdown: List[str] = []
        sheets_data: List[Dict[str, Any]] = []

        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_name = sheet.name

            max_row = min(sheet.nrows, self.max_rows)
            max_col = min(sheet.ncols, self.max_cols)

            if max_row == 0 or max_col == 0:
                continue

            # 读取所有行（保留原始值和类型，用于 normalize_cell_value）
            raw_rows: List[List[Any]] = []
            for row_idx in range(max_row):
                row_data = []
                for col_idx in range(max_col):
                    cell = sheet.cell(row_idx, col_idx)
                    # 传入 xlrd_wb 和 xlrd_cell_type 以正确处理日期序列号
                    normalized = normalize_cell_value(cell.value, xlrd_wb=wb, xlrd_cell_type=cell.ctype)
                    row_data.append(normalized)
                raw_rows.append(row_data)

            if not raw_rows:
                continue

            header_row_idx = find_header_row(raw_rows, self.scan_header_rows)
            header = [raw_rows[header_row_idx][i] or f"列{i+1}" for i in range(max_col)]
            data_raw = raw_rows[header_row_idx + 1:]
            str_rows = [r for r in data_raw if any(v for v in r)]

            if not str_rows:
                continue

            column_types = infer_column_types(header, str_rows)
            sheet_md = rows_to_markdown(sheet_name, header, str_rows, row_start=1)
            if sheet_md:
                sheets_markdown.append(sheet_md)

            sheets_data.append({
                "sheet_name": sheet_name,
                "header": header,
                "header_row_index": header_row_idx,
                "rows": str_rows,
                "column_types": column_types,
                "has_merged_cells": False,  # xlrd 不支持合并单元格回填
                "formula_none_cells": 0,
            })

            # on_demand 模式下及时释放 Sheet 内存
            wb.unload_sheet(sheet_idx)

        markdown = "\n\n---\n\n".join(sheets_markdown)
        metadata = {
            "parse_method": "excel",
            "parser": "xlrd",
            "format": "xls",
            "sheet_count": len(sheets_data),
            "sheets": [{"name": s["sheet_name"], "rows": len(s["rows"]), "cols": len(s["header"])} for s in sheets_data],
            "sheets_data": sheets_data,
        }

        logger.info(f"[ExcelParser] xls 解析完成，有效 Sheet 数: {len(sheets_data)}")
        return markdown, metadata
