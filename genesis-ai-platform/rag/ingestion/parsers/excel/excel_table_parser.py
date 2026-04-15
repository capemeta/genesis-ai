"""
Excel 表格模式专用解析器

与 ExcelParser（通用模式）的区别：
- 不输出 Markdown text，直接输出结构化 table_rows 列表
- 专为"一行 = 一条记录"的场景设计
- 支持列类型推断、field_map 生成
- field_map 在 metadata 中返回，不在 parser 内回写 KB（层次隔离，由 Service 层决定是否回写）

输出 metadata 结构：
    {
        "parse_method": "excel_table",
        "parser": "openpyxl" | "xlrd" | "pandas_fallback",
        "format": "xlsx" | "xls",
        "table_rows": [
            {
                "sheet_name": str,
                "row_index": int,      # 1-based，相对于数据区域（不含表头）
                "header": list[str],
                "values": list[str],
                "column_types": dict[str, str],
            },
            ...
        ],
        "sheets": [
            {
                "sheet_name": str,
                "header": list[str],
                "row_count": int,
                "column_types": dict[str, str],
            },
        ],
        "field_map": {"地区": "text", "事项名称": "text", ...},  # 列名→类型，供 Service 层回写 KB
    }
"""

import logging
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from .excel_parser_utils import (
    backfill_merged_cells,
    find_header_row,
    infer_column_types,
    normalize_cell_value,
)

logger = logging.getLogger(__name__)

FULL_LOAD_SIZE_LIMIT = 50 * 1024 * 1024  # 50MB


@dataclass
class ExcelRowData:
    """表格模式单行数据。"""
    sheet_name: str
    row_index: int                          # 1-based，相对于数据区域
    header: List[str]
    values: List[str]
    column_types: Dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "row_index": self.row_index,
            "header": self.header,
            "values": self.values,
            "column_types": self.column_types,
        }


class ExcelTableParser:
    """
    Excel 表格模式专用解析器。

    输出结构化行数据（table_rows），供 ExcelTableChunker 消费。
    不输出 Markdown text（表格模式无需整体 Markdown）。
    """

    def __init__(
        self,
        max_rows: int = 100000,
        max_cols: int = 200,
        scan_header_rows: int = 10,
        include_hidden_sheets: bool = False,
        row_from: int = 0,
        row_to: Optional[int] = None,
    ):
        """
        Args:
            max_rows: 每个工作表最大行数
            max_cols: 每个工作表最大列数
            scan_header_rows: 智能表头检测扫描行数
            include_hidden_sheets: 是否处理隐藏工作表
            row_from: 读取数据行的起始行（0-based，相对于数据区域，用于大表分片）
            row_to: 读取数据行的结束行（不含，None 表示读到末尾）
        """
        self.max_rows = max_rows
        self.max_cols = max_cols
        self.scan_header_rows = scan_header_rows
        self.include_hidden_sheets = include_hidden_sheets
        self.row_from = row_from
        self.row_to = row_to

    def parse(
        self,
        file_buffer: bytes,
        file_extension: str = ".xlsx",
    ) -> Tuple[str, Dict[str, Any]]:
        """
        解析 Excel 文件（表格模式）。

        Args:
            file_buffer: Excel 文件字节内容
            file_extension: 文件扩展名

        Returns:
            ("", metadata)：text 始终为空字符串，metadata 含 table_rows
        """
        logger.info(f"[ExcelTableParser] 开始解析（表格模式），格式: {file_extension}, 大小: {len(file_buffer)} bytes")

        if file_extension.lower() == ".xls":
            metadata = self._parse_xls(file_buffer)
        else:
            metadata = self._parse_xlsx(file_buffer)

        return "", metadata

    # ------------------------------------------------------------------ #
    # .xlsx 解析
    # ------------------------------------------------------------------ #

    def _parse_xlsx(self, file_buffer: bytes) -> Dict[str, Any]:
        """解析 .xlsx，openpyxl 失败时降级 pandas。"""
        try:
            return self._parse_xlsx_openpyxl(file_buffer)
        except Exception as e:
            logger.warning(f"[ExcelTableParser] openpyxl 解析失败，降级 pandas: {e}")
            return self._parse_xlsx_pandas(file_buffer)

    def _parse_xlsx_openpyxl(self, file_buffer: bytes) -> Dict[str, Any]:
        try:
            import openpyxl  # type: ignore[import-untyped]
        except ImportError:
            raise RuntimeError("openpyxl 未安装，请执行: pip install openpyxl")

        file_size = len(file_buffer)
        use_read_only = file_size > FULL_LOAD_SIZE_LIMIT

        if use_read_only:
            wb = openpyxl.load_workbook(BytesIO(file_buffer), read_only=True, data_only=True, keep_vba=False)
        else:
            wb = openpyxl.load_workbook(BytesIO(file_buffer), read_only=False, data_only=True, keep_vba=False)

        all_table_rows: List[Dict[str, Any]] = []
        sheets_info: List[Dict[str, Any]] = []
        field_map: Dict[str, str] = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            if not self.include_hidden_sheets:
                state = getattr(sheet, "sheet_state", None)
                if state == "hidden":
                    logger.debug(f"[ExcelTableParser] 跳过隐藏 Sheet: {sheet_name}")
                    continue

            if not use_read_only:
                backfill_merged_cells(sheet)

            # 读取全部行（用于表头检测，实际数据按 row_from/row_to 切片）
            raw_rows: List[List[Any]] = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx >= self.max_rows + self.scan_header_rows:
                    break
                raw_rows.append(list(row[:self.max_cols]))

            if not raw_rows:
                continue

            header_row_idx = find_header_row(raw_rows, self.scan_header_rows)
            header_raw = raw_rows[header_row_idx]
            header = [normalize_cell_value(v) or f"列{i+1}" for i, v in enumerate(header_raw[:self.max_cols])]

            data_raw_all = raw_rows[header_row_idx + 1:]

            # 大表分片：仅读取 [row_from, row_to) 范围内的数据行
            data_raw = data_raw_all[self.row_from: self.row_to]

            # 规范化
            str_rows: List[List[str]] = []
            for row in data_raw:
                str_row = [normalize_cell_value(v) for v in row[:len(header)]]
                str_row += [""] * max(0, len(header) - len(str_row))
                str_rows.append(str_row)
            str_rows = [r for r in str_rows if any(v for v in r)]

            if not str_rows:
                continue

            # 列类型推断（用全量数据行的前 100 行）
            column_types = infer_column_types(header, str_rows[:100])

            # 生成 ExcelRowData
            for local_idx, excel_row in enumerate(str_rows):
                global_row_index = self.row_from + local_idx + 1  # 1-based，全局行号
                row_data = ExcelRowData(
                    sheet_name=sheet_name,
                    row_index=global_row_index,
                    header=header,
                    values=excel_row,
                    column_types=column_types,
                )
                all_table_rows.append(row_data.to_dict())

            sheets_info.append({
                "sheet_name": sheet_name,
                "header": header,
                "header_row_number": header_row_idx + 1,
                "row_count": len(data_raw_all),  # 全量数据行数（非分片范围）
                "column_types": column_types,
            })

            # field_map：列名 → 类型（取第一个非空 Sheet 的列定义）
            if not field_map:
                field_map = column_types

        parser_name = "openpyxl" if not use_read_only else "openpyxl_readonly"
        return self._build_metadata(all_table_rows, sheets_info, field_map, parser_name, "xlsx")

    def _parse_xlsx_pandas(self, file_buffer: bytes) -> Dict[str, Any]:
        try:
            import pandas as pd  # type: ignore[import-untyped]
        except ImportError:
            raise RuntimeError("pandas 未安装，请执行: pip install pandas openpyxl")

        xls = pd.ExcelFile(BytesIO(file_buffer), engine="openpyxl")
        all_table_rows: List[Dict[str, Any]] = []
        sheets_info: List[Dict[str, Any]] = []
        field_map: Dict[str, str] = {}

        for sheet_name in xls.sheet_names:
            df = xls.parse(sheet_name, header=None, dtype=str)
            if df.empty:
                continue

            raw_rows = df.values.tolist()
            header_row_idx = find_header_row(raw_rows, self.scan_header_rows)
            header_raw = raw_rows[header_row_idx]
            header = [str(v).strip() if v and str(v).strip() not in ("nan", "None") else f"列{i+1}"
                      for i, v in enumerate(header_raw[:self.max_cols])]

            data_raw_all = raw_rows[header_row_idx + 1:]
            data_raw = data_raw_all[self.row_from: self.row_to]

            str_rows: List[List[str]] = []
            for row in data_raw:
                str_row = [str(v).strip() if v and str(v).strip() not in ("nan", "None") else ""
                           for v in row[:len(header)]]
                str_row += [""] * max(0, len(header) - len(str_row))
                str_rows.append(str_row)
            str_rows = [r for r in str_rows if any(v for v in r)]

            if not str_rows:
                continue

            column_types = infer_column_types(header, str_rows[:100])

            for local_idx, excel_row in enumerate(str_rows):
                global_row_index = self.row_from + local_idx + 1
                row_data = ExcelRowData(
                    sheet_name=sheet_name,
                    row_index=global_row_index,
                    header=header,
                    values=excel_row,
                    column_types=column_types,
                )
                all_table_rows.append(row_data.to_dict())

            sheets_info.append({
                "sheet_name": sheet_name,
                "header": header,
                "header_row_number": header_row_idx + 1,
                "row_count": len(data_raw_all),
                "column_types": column_types,
            })

            if not field_map:
                field_map = column_types

        return self._build_metadata(all_table_rows, sheets_info, field_map, "pandas_fallback", "xlsx")

    # ------------------------------------------------------------------ #
    # .xls 解析
    # ------------------------------------------------------------------ #

    def _parse_xls(self, file_buffer: bytes) -> Dict[str, Any]:
        try:
            import xlrd  # type: ignore[import-untyped]
        except ImportError:
            raise RuntimeError("xlrd 未安装，请执行: pip install xlrd")

        wb = xlrd.open_workbook(file_contents=file_buffer, on_demand=True)
        all_table_rows: List[Dict[str, Any]] = []
        sheets_info: List[Dict[str, Any]] = []
        field_map: Dict[str, str] = {}

        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_name = sheet.name

            max_row = min(sheet.nrows, self.max_rows + self.scan_header_rows)
            max_col = min(sheet.ncols, self.max_cols)

            if max_row == 0:
                continue

            # xlrd 需要传入 wb 和 cell_type 以正确处理日期
            raw_rows: List[List[str]] = []
            for row_idx in range(max_row):
                cell_values: List[str] = []
                for col_idx in range(max_col):
                    cell = sheet.cell(row_idx, col_idx)
                    cell_values.append(normalize_cell_value(cell.value, xlrd_wb=wb, xlrd_cell_type=cell.ctype))
                raw_rows.append(cell_values)

            header_row_idx = find_header_row(raw_rows, self.scan_header_rows)
            header = [raw_rows[header_row_idx][i] or f"列{i+1}" for i in range(max_col)]

            data_raw_all = raw_rows[header_row_idx + 1:]
            data_raw = data_raw_all[self.row_from: self.row_to]
            str_rows: List[List[str]] = [r for r in data_raw if any(v for v in r)]

            if not str_rows:
                wb.unload_sheet(sheet_idx)
                continue

            column_types = infer_column_types(header, str_rows[:100])

            for local_idx, str_row in enumerate(str_rows):
                global_row_index = self.row_from + local_idx + 1
                row_data = ExcelRowData(
                    sheet_name=sheet_name,
                    row_index=global_row_index,
                    header=header,
                    values=str_row,
                    column_types=column_types,
                )
                all_table_rows.append(row_data.to_dict())

            sheets_info.append({
                "sheet_name": sheet_name,
                "header": header,
                "header_row_number": header_row_idx + 1,
                "row_count": len(data_raw_all),
                "column_types": column_types,
            })

            if not field_map:
                field_map = column_types

            wb.unload_sheet(sheet_idx)

        return self._build_metadata(all_table_rows, sheets_info, field_map, "xlrd", "xls")

    # ------------------------------------------------------------------ #
    # 构建 metadata
    # ------------------------------------------------------------------ #

    def _build_metadata(
        self,
        table_rows: List[Dict[str, Any]],
        sheets_info: List[Dict[str, Any]],
        field_map: Dict[str, str],
        parser_name: str,
        fmt: str,
    ) -> Dict[str, Any]:
        logger.info(f"[ExcelTableParser] 解析完成，共 {len(table_rows)} 行，{len(sheets_info)} 个 Sheet")
        return {
            "parse_method": "excel_table",
            "parser": parser_name,
            "format": fmt,
            "sheet_count": len(sheets_info),
            "sheets": sheets_info,
            "table_rows": table_rows,
            # field_map 由 Service 层决定是否写回 KB 配置
            "field_map": field_map,
        }
